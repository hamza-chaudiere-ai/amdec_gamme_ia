#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur d'analyses AMDEC - VERSION REGROUPEMENT COMPOSANT+SOUS-COMPOSANT
✅ REGROUPEMENT INTELLIGENT: Une seule ligne par composant+sous-composant
✅ FUSION DES CAUSES: Concaténation de toutes les causes différentes
✅ CALCUL GLOBAL: Fréquence basée sur le nombre total d'occurrences
✅ MODES FUSIONNÉS: Combination intelligente des modes de défaillance
"""

import pandas as pd
import numpy as np
import os
import logging
import re
from datetime import datetime
from typing import Dict, List, Optional, Union, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .utils import (
    format_component_display,
    format_subcomponent_display,
    calculate_criticality,
    get_criticality_color,
    generate_timestamp,
    ComponentConfig
)

logger = logging.getLogger(__name__)

class AMDECGenerator:
    """
    ✅ GÉNÉRATEUR AMDEC avec REGROUPEMENT par COMPOSANT+SOUS-COMPOSANT
    
    Fonctionnalités spéciales :
    - Regroupement unique par composant + sous-composant (ignorer causes)
    - Fusion intelligente des causes multiples
    - Calcul de fréquence globale selon grille officielle
    - Concaténation des modes de défaillance
    """
    
    def __init__(self, data: Optional[pd.DataFrame] = None):
        """
        Initialise le générateur AMDEC avec regroupement intelligent
        
        Args:
            data: DataFrame avec données d'historique (optionnel)
        """
        self.data = data
        self.original_data = None
        self.amdec_data = []
        self.amdec_df = None
        self.output_path = None
        self.grouping_stats = {}
        
        # Chargement des modèles de référence
        self.reference_models = self._load_reference_models()
        
        # Base de connaissances pour la génération intelligente
        self.knowledge_base = self._build_knowledge_base()
    
    def generate(self) -> pd.DataFrame:
        """
        ✅ GÉNÉRATION AMDEC avec REGROUPEMENT COMPOSANT+SOUS-COMPOSANT
        """
        try:
            logger.info("🔄 Début génération AMDEC avec regroupement composant+sous-composant")
            
            if self.data is not None and not self.data.empty:
                # ✅ TRAITEMENT avec regroupement par composant+sous-composant
                self._process_historical_data_with_component_grouping()
            else:
                # Génération à partir des modèles de référence
                self._generate_from_models()
            
            # Post-traitement et optimisation
            self._optimize_amdec()
            
            # ✅ Création DataFrame PROPRE (sans colonnes supplémentaires)
            self._create_clean_dataframe()
            
            logger.info(f"✅ AMDEC générée: {len(self.amdec_data)} entrées consolidées par composant+sous-composant")
            return self.amdec_df
            
        except Exception as e:
            logger.error(f"❌ Erreur lors de la génération AMDEC: {e}")
            raise
    
    def _process_historical_data_with_component_grouping(self):
        """
        ✅ TRAITEMENT SPÉCIAL avec regroupement par COMPOSANT+SOUS-COMPOSANT uniquement
        
        Pipeline :
        1. Validation des données d'entrée
        2. Regroupement par (composant, sous_composant) SEULEMENT
        3. Fusion de toutes les causes pour chaque groupe
        4. Calcul de fréquence globale selon grille officielle
        5. Génération des entrées AMDEC consolidées
        """
        logger.info("🔄 Traitement avec regroupement COMPOSANT+SOUS-COMPOSANT")
        
        # Sauvegarder les données originales
        self.original_data = self.data.copy()
        
        # ✅ ÉTAPE 1: Validation des données d'entrée
        required_columns = ['composant', 'sous_composant']
        missing_columns = [col for col in required_columns if col not in self.data.columns]
        
        if missing_columns:
            raise ValueError(f"❌ Colonnes manquantes pour regroupement: {missing_columns}")
        
        # S'assurer que les colonnes optionnelles existent
        if 'cause' not in self.data.columns:
            self.data['cause'] = 'non_specifie'
        if 'duree' not in self.data.columns:
            self.data['duree'] = 1.0
        
        logger.info(f"📊 Données d'entrée: {len(self.data)} lignes")
        
        # ✅ ÉTAPE 2: Regroupement EXCLUSIVEMENT par composant+sous-composant
        self._perform_component_subcomponent_grouping()
        
        logger.info(f"✅ Regroupement terminé: {len(self.amdec_data)} entrées AMDEC uniques")
    
    def _perform_component_subcomponent_grouping(self):
        """
        ✅ REGROUPEMENT PRINCIPAL par COMPOSANT+SOUS-COMPOSANT uniquement
        
        Toutes les causes différentes sont fusionnées pour un même composant+sous-composant
        """
        logger.info("📊 Regroupement par COMPOSANT + SOUS-COMPOSANT...")
        
        # ✅ Regroupement sur SEULEMENT (composant, sous-composant)
        grouping_cols = ['composant', 'sous_composant']
        
        logger.info(f"📊 Regroupement sur: {grouping_cols}")
        logger.info(f"📊 Données avant regroupement: {len(self.data)} lignes")
        
        # Effectuer le regroupement avec agrégations intelligentes
        try:
            grouped = self.data.groupby(grouping_cols, dropna=False).agg({
                'cause': lambda x: list(x),  # Toutes les causes
                'duree': ['count', 'mean', 'sum', 'max', 'min'],  # Statistiques durées
                'date_arret': 'count' if 'date_arret' in self.data.columns else lambda x: len(x),
                'line_id': 'count' if 'line_id' in self.data.columns else lambda x: len(x)
            }).reset_index()
            
            # Aplatir les colonnes multi-niveaux
            grouped.columns = [
                'composant', 'sous_composant',
                'causes_list', 
                'nb_occurrences', 'duree_moyenne', 'duree_totale', 'duree_max', 'duree_min',
                'date_count', 'line_count'
            ]
            
        except Exception as e:
            logger.error(f"❌ Erreur regroupement: {e}")
            grouped = self._fallback_simple_grouping()
        
        logger.info(f"📊 Regroupement effectué: {len(self.data)} lignes → {len(grouped)} groupes uniques")
        
        # ✅ ÉTAPE 3: Traiter chaque groupe et créer entrée AMDEC
        for _, group in grouped.iterrows():
            try:
                self._create_amdec_entry_from_group(group)
            except Exception as e:
                logger.warning(f"⚠️ Erreur traitement groupe: {e}")
                continue
        
        # Statistiques de regroupement
        self.grouping_stats = {
            'original_lines': len(self.data),
            'grouped_entries': len(grouped),
            'amdec_entries_created': len(self.amdec_data),
            'compression_ratio': len(grouped) / len(self.data) if len(self.data) > 0 else 0
        }
        
        logger.info(f"📊 Statistiques regroupement: {len(self.data)} → {len(grouped)} groupes → {len(self.amdec_data)} AMDEC")
    
    def _create_amdec_entry_from_group(self, group_data: pd.Series):
        """
        ✅ CRÉATION d'une entrée AMDEC à partir d'un groupe consolidé
        
        Args:
            group_data: Données du groupe (composant+sous-composant avec toutes ses causes)
        """
        try:
            component = str(group_data['composant'])
            subcomponent = str(group_data['sous_composant'])
            
            # Valider les données de base
            if any(val in ['nan', 'None', ''] for val in [component, subcomponent]):
                logger.debug(f"⚠️ Données incomplètes ignorées: {component}-{subcomponent}")
                return
            
            # ✅ TRAITEMENT des causes multiples
            causes_list = group_data['causes_list']
            if isinstance(causes_list, list):
                unique_causes = list(set([str(c) for c in causes_list if str(c) not in ['nan', 'None', 'non_specifie']]))
            else:
                unique_causes = [str(causes_list)]
            
            # Supprimer les causes vides
            unique_causes = [c for c in unique_causes if c.strip() != '']
            if not unique_causes:
                unique_causes = ['non_specifie']
            
            # ✅ FUSION INTELLIGENTE des causes
            if len(unique_causes) == 1:
                consolidated_cause = unique_causes[0]
            elif len(unique_causes) <= 3:
                consolidated_cause = ' + '.join(unique_causes)
            else:
                # Si trop de causes, prendre les 3 principales + "autres"
                main_causes = unique_causes[:3]
                consolidated_cause = ' + '.join(main_causes) + f' (+{len(unique_causes)-3} autres)'
            
            # ✅ CALCUL de la fréquence OFFICIELLE selon grille
            nb_occurrences = int(group_data['nb_occurrences'])
            frequency = self._calculate_frequency_OFFICIAL_GRID(nb_occurrences)
            
            # ✅ CALCUL de la gravité améliorée
            if pd.notna(group_data['duree_moyenne']):
                duree_moyenne = float(group_data['duree_moyenne'])
                duree_max = float(group_data.get('duree_max', duree_moyenne))
                gravity = self._calculate_gravity_ENHANCED(duree_moyenne, duree_max, nb_occurrences)
            else:
                gravity = self._estimate_gravity_from_causes(unique_causes)
            
            # ✅ CALCUL de la détection intelligente
            detection = self._calculate_detection_from_causes_SMART(unique_causes)
            
            # ✅ Criticité = F × G × D
            criticality = calculate_criticality(frequency, gravity, detection)
            
            # ✅ GÉNÉRATION des autres attributs intelligents
            failure_modes = self._determine_failure_modes_MULTIPLE(unique_causes, subcomponent, nb_occurrences)
            consolidated_failure_mode = self._consolidate_failure_modes(failure_modes)
            
            effect = self._determine_effect_COMPREHENSIVE(consolidated_failure_mode, gravity, nb_occurrences, len(unique_causes))
            function = self._determine_function_ENHANCED(component, subcomponent)
            actions = self._generate_corrective_actions_COMPREHENSIVE(unique_causes, criticality, nb_occurrences)
            
            # ✅ CRÉER l'entrée AMDEC FINALE
            entry = {
                'Composant': format_component_display(component),
                'Sous-composant': format_subcomponent_display(subcomponent),
                'Fonction': function,
                'Mode de Défaillance': consolidated_failure_mode,
                'Cause': consolidated_cause,
                'Effet': effect,
                'F': frequency,  # ✅ Fréquence selon GRILLE OFFICIELLE
                'G': gravity,
                'D': detection,
                'C': criticality,  # ✅ Criticité recalculée OFFICIELLEMENT
                'Actions Correctives': actions
            }
            
            self.amdec_data.append(entry)
            
            logger.debug(f"✅ Entrée AMDEC créée: {component}-{subcomponent} → C={criticality} (F={frequency} pour {nb_occurrences} occurrences, {len(unique_causes)} causes)")
            
        except Exception as e:
            logger.error(f"❌ Erreur création entrée AMDEC: {e}")
    
    def _calculate_frequency_OFFICIAL_GRID(self, nb_occurrences: int) -> int:
        """
        ✅ GRILLE OFFICIELLE: Calcul EXACT de la fréquence selon grille demandée
        
        GRILLE OFFICIELLE:
        - 1 occurrence = F=1 (Très faible - rare, <1/an)
        - 2 à 3 occurrences = F=2 (Faible - possible, <1/trimestre)
        - 4 à 7 occurrences = F=3 (Moyenne - fréquente, <1/semaine)
        - 8+ occurrences = F=4 (Forte - plusieurs fois/semaine)
        
        Args:
            nb_occurrences: Nombre TOTAL d'occurrences pour ce composant+sous-composant
            
        Returns:
            Fréquence F selon grille officielle
        """
        if nb_occurrences == 1:
            return 1  # Très faible
        elif 2 <= nb_occurrences <= 3:
            return 2  # Faible
        elif 4 <= nb_occurrences <= 7:
            return 3  # Moyenne
        else:  # 8 et plus
            return 4  # Forte
    
    def _calculate_gravity_ENHANCED(self, duree_moyenne: float, duree_max: float, nb_occurrences: int) -> int:
        """
        ✅ CALCUL AMÉLIORÉ de la gravité G
        
        Args:
            duree_moyenne: Durée moyenne des arrêts
            duree_max: Durée maximale observée
            nb_occurrences: Nombre d'occurrences
            
        Returns:
            Gravité G (1-5)
        """
        # Utiliser la durée la plus représentative selon le nombre d'occurrences
        if nb_occurrences >= 5:
            duree_reference = duree_moyenne  # Moyenne fiable
        elif nb_occurrences >= 3:
            duree_reference = (duree_moyenne + duree_max) / 2  # Mix moyenne/max
        else:
            duree_reference = duree_max  # Prendre le pire cas
        
        # Ajustement selon la récurrence (plus d'occurrences = impact plus grave)
        if nb_occurrences >= 8:
            duree_reference *= 1.3  # Aggravation significative
        elif nb_occurrences >= 5:
            duree_reference *= 1.1  # Aggravation modérée
        
        # Barème de gravité industriel (en heures)
        if duree_reference < 0.5:
            return 1  # Négligeable (< 30 min)
        elif duree_reference < 2:
            return 2  # Mineure (< 2h)
        elif duree_reference < 8:
            return 3  # Modérée (< 8h)
        elif duree_reference < 24:
            return 4  # Majeure (< 1 jour)
        else:
            return 5  # Catastrophique (> 1 jour)
    
    def _estimate_gravity_from_causes(self, causes: List[str]) -> int:
        """
        ✅ ESTIMATION de gravité basée sur les types de causes
        
        Args:
            causes: Liste des causes uniques
            
        Returns:
            Gravité estimée (1-5)
        """
        max_gravity = 2  # Gravité de base
        
        # Mapping des gravités par type de cause
        gravity_mapping = {
            'surchauffe': 4,     # Grave car peut causer rupture
            'fissure': 4,        # Grave car peut évoluer vers rupture
            'rupture': 5,        # Catastrophique
            'explosion': 5,      # Catastrophique
            'percement': 4,      # Grave
            'corrosion': 3,      # Modérée mais progressive
            'erosion': 3,        # Modérée mais progressive
            'fatigue': 3,        # Modérée mais répétitive
            'usure': 2,          # Mineure si détectée tôt
            'encrassement': 2,   # Mineure, réversible
            'vibration': 2,      # Mineure mais peut évoluer
            'fuite': 3           # Modérée selon l'ampleur
        }
        
        # Prendre la gravité maximale parmi toutes les causes
        for cause in causes:
            cause_lower = cause.lower()
            for cause_key, gravity in gravity_mapping.items():
                if cause_key in cause_lower:
                    max_gravity = max(max_gravity, gravity)
        
        # Bonus si multiples causes (combinaison = plus grave)
        if len(causes) >= 3:
            max_gravity = min(max_gravity + 1, 5)
        
        return max_gravity
    
    def _calculate_detection_from_causes_SMART(self, causes: List[str]) -> int:
        """
        ✅ CALCUL INTELLIGENT de la détection D basé sur TOUTES les causes
        
        Args:
            causes: Liste des causes uniques
            
        Returns:
            Détection D (1-4) - Prend la MEILLEURE détection (plus facile)
        """
        min_detection = 4  # Commencer par le pire (très difficile)
        
        detection_mapping = {
            # Très facile à détecter (D=1)
            'fuite': 1, 'percement': 1, 'rupture': 1, 'explosion': 1, 'surchauffe': 1,
            'vibration': 1, 'bruit': 1, 'temperature': 1, 'alerte': 1,
            
            # Facile à détecter (D=2)
            'corrosion': 2, 'erosion': 2, 'encrassement': 2, 'obstruction': 2,
            'deformation': 2, 'usure': 2, 'rouille': 2,
            
            # Difficile à détecter (D=3)
            'fissure': 3, 'fatigue': 3, 'microfissure': 3, 'contrainte': 3,
            'dilatation': 3, 'fluage': 3,
            
            # Très difficile à détecter (D=4)
            'defaut_interne': 4, 'vieillissement': 4, 'degradation_lente': 4,
            'contamination': 4, 'impurete': 4, 'non_specifie': 4
        }
        
        # Prendre la MEILLEURE détection (plus facile = valeur plus faible)
        for cause in causes:
            cause_lower = cause.lower()
            for cause_key, detection_value in detection_mapping.items():
                if cause_key in cause_lower:
                    min_detection = min(min_detection, detection_value)
        
        return min_detection
    
    def _determine_failure_modes_MULTIPLE(self, causes: List[str], subcomponent: str, nb_occurrences: int) -> List[str]:
        """
        ✅ DÉTERMINATION de modes de défaillance MULTIPLES
        
        Args:
            causes: Liste des causes
            subcomponent: Sous-composant
            nb_occurrences: Nombre d'occurrences
            
        Returns:
            Liste des modes de défaillance
        """
        modes = []
        subcomp_lower = subcomponent.lower()
        
        smart_mapping = {
            'corrosion': {
                'epingle': 'Corrosion externe généralisée',
                'collecteur': 'Corrosion interne par chimie eau',
                'tube': 'Corrosion côté feu/fumées',
                'branches': 'Corrosion sous contrainte',
                'default': 'Corrosion progressive'
            },
            'erosion': {
                'epingle': 'Érosion par cendres volantes',
                'collecteur': 'Érosion par circulation fluide',
                'tube': 'Érosion par particules en suspension',
                'branches': 'Érosion par vitesse excessive',
                'default': 'Érosion mécanique'
            },
            'fissure': {
                'epingle': 'Fissuration thermique cyclique',
                'collecteur': 'Fissures intergranulaires',
                'tube': 'Fatigue thermomécanique',
                'branches': 'Fissuration par contrainte',
                'default': 'Fissuration progressive'
            },
            'fatigue': {
                'epingle': 'Fatigue thermique cyclique',
                'collecteur': 'Fatigue mécanique',
                'tube': 'Fatigue par cycles pression',
                'branches': 'Fatigue vibratoire',
                'default': 'Fatigue cyclique'
            },
            'surchauffe': {
                'epingle': 'Surchauffe locale critique',
                'collecteur': 'Surchauffe par stagnation',
                'tube': 'Surchauffe long terme (fluage)',
                'branches': 'Surchauffe par déséquilibre',
                'default': 'Surchauffe opérationnelle'
            }
        }
        
        # Déterminer le type de sous-composant
        comp_type = 'default'
        for key in ['epingle', 'collecteur', 'tube', 'branches']:
            if key in subcomp_lower:
                comp_type = key
                break
        
        # Générer un mode pour chaque cause
        for cause in causes:
            cause_lower = cause.lower()
            mode_found = False
            
            for cause_key, modes_dict in smart_mapping.items():
                if cause_key in cause_lower:
                    base_mode = modes_dict.get(comp_type, modes_dict['default'])
                    modes.append(base_mode)
                    mode_found = True
                    break
            
            if not mode_found:
                modes.append(f'Défaillance par {cause}')
        
        return modes
    
    def _consolidate_failure_modes(self, modes: List[str]) -> str:
        """
        ✅ CONSOLIDATION intelligente des modes de défaillance multiples
        
        Args:
            modes: Liste des modes de défaillance
            
        Returns:
            Mode consolidé
        """
        if not modes:
            return 'Mode de défaillance non spécifié'
        
        if len(modes) == 1:
            return modes[0]
        elif len(modes) <= 3:
            return ' + '.join(modes)
        else:
            # Si trop de modes, grouper par famille
            mode_families = {}
            for mode in modes:
                if 'corrosion' in mode.lower():
                    mode_families.setdefault('Corrosion', []).append(mode)
                elif 'fissure' in mode.lower() or 'fatigue' in mode.lower():
                    mode_families.setdefault('Fatigue/Fissuration', []).append(mode)
                elif 'erosion' in mode.lower() or 'usure' in mode.lower():
                    mode_families.setdefault('Érosion/Usure', []).append(mode)
                elif 'surchauffe' in mode.lower():
                    mode_families.setdefault('Surchauffe', []).append(mode)
                else:
                    mode_families.setdefault('Autres', []).append(mode)
            
            if len(mode_families) <= 3:
                return ' + '.join(mode_families.keys())
            else:
                main_families = list(mode_families.keys())[:2]
                return ' + '.join(main_families) + f' (+{len(mode_families)-2} autres)'
    
    def _determine_effect_COMPREHENSIVE(self, failure_mode: str, gravity: int, nb_occurrences: int, nb_causes: int) -> str:
        """
        ✅ EFFET COMPLET basé sur mode, gravité, récurrence ET multiplicité des causes
        
        Args:
            failure_mode: Mode de défaillance consolidé
            gravity: Gravité
            nb_occurrences: Nombre d'occurrences
            nb_causes: Nombre de causes différentes
            
        Returns:
            Effet détaillé
        """
        mode_lower = failure_mode.lower()
        
        # Effet de base selon le mode principal
        if 'corrosion' in mode_lower:
            if gravity <= 2:
                base_effect = 'Amincissement local des parois'
            elif gravity <= 3:
                base_effect = 'Perte de matière significative'
            else:
                base_effect = 'Perforation et fuite majeure'
                
        elif 'fissure' in mode_lower or 'fatigue' in mode_lower:
            if gravity <= 2:
                base_effect = 'Microfissures superficielles'
            elif gravity <= 3:
                base_effect = 'Fissures traversantes locales'
            else:
                base_effect = 'Rupture catastrophique'
                
        elif 'erosion' in mode_lower:
            if gravity <= 2:
                base_effect = 'Usure superficielle progressive'
            elif gravity <= 3:
                base_effect = 'Amincissement accéléré'
            else:
                base_effect = 'Perforation par érosion'
                
        elif 'surchauffe' in mode_lower:
            if gravity <= 2:
                base_effect = 'Déformation plastique mineure'
            elif gravity <= 3:
                base_effect = 'Fluage et déformation permanente'
            else:
                base_effect = 'Rupture ductile immédiate'
        else:
            if gravity <= 2:
                base_effect = 'Impact mineur sur performance'
            elif gravity <= 3:
                base_effect = 'Dégradation notable performance'
            else:
                base_effect = 'Arrêt imprévu et coûteux'
        
        # Ajouts selon la récurrence
        if nb_occurrences >= 8:
            base_effect = f"{base_effect} + Récurrence CRITIQUE ({nb_occurrences} fois)"
        elif nb_occurrences >= 5:
            base_effect = f"{base_effect} + Récurrence préoccupante ({nb_occurrences} fois)"
        elif nb_occurrences >= 3:
            base_effect = f"{base_effect} + Tendance répétitive"
        
        # Ajouts selon la multiplicité des causes
        if nb_causes >= 4:
            base_effect = f"{base_effect} + Défaillance MULTI-CAUSES ({nb_causes} causes)"
        elif nb_causes >= 2:
            base_effect = f"{base_effect} + Causes multiples"
        
        return base_effect
    
    def _determine_function_ENHANCED(self, component: str, subcomponent: str) -> str:
        """
        ✅ FONCTION précise selon l'expertise industrielle
        
        Args:
            component: Composant
            subcomponent: Sous-composant
            
        Returns:
            Fonction détaillée
        """
        comp_lower = component.lower()
        subcomp_lower = subcomponent.lower()
        
        function_mapping = {
            'economiseur': {
                'epingle': 'Récupération chaleur fumées → préchauffage eau alimentation',
                'collecteur_entree': 'Distribution eau alimentation vers nappes',
                'collecteur_sortie': 'Collecte eau préchauffée vers ballon',
                'tubes_suspension': 'Support mécanique + guidage dilatation thermique'
            },
            'surchauffeur': {
                'epingle': 'Surchauffe vapeur saturée → vapeur haute température service',
                'tube_porteur': 'Résistance pression service + transfert thermique optimal',
                'branches_entree': 'Distribution vapeur saturée vers nappes surchauffe',
                'branches_sortie': 'Collecte vapeur surchauffée vers collecteur principal',
                'collecteur_entree': 'Distribution vapeur saturée depuis ballon',
                'collecteur_sortie': 'Collecte vapeur surchauffée finale vers turbine'
            },
            'rechauffeur': {
                'collecteur_entree': 'Distribution vapeur partiellement détendue',
                'collecteur_sortie': 'Collecte vapeur réchauffée vers turbine BP',
                'tubes_suspension': 'Support structural + résistance cycles thermiques',
                'tube_porteur': 'Support mécanique résistant haute pression + température',
                'branches_entree': 'Distribution vapeur vers échangeur réchauffage',
                'branches_sortie': 'Évacuation vapeur réchauffée optimisée vers turbine'
            }
        }
        
        # Identifier le type de composant principal
        comp_type = None
        for key in ['economiseur', 'surchauffeur', 'rechauffeur']:
            if key in comp_lower:
                comp_type = key
                break
        
        if comp_type and comp_type in function_mapping:
            for subcomp_key, function in function_mapping[comp_type].items():
                if subcomp_key in subcomp_lower or any(part in subcomp_lower for part in subcomp_key.split('_')):
                    return function
        
        return 'Fonction de transfert thermique et résistance mécanique sous pression'
    
    def _generate_corrective_actions_COMPREHENSIVE(self, causes: List[str], criticality: int, nb_occurrences: int) -> str:
        """
        ✅ ACTIONS CORRECTIVES COMPLÈTES pour causes multiples
        
        Args:
            causes: Liste des causes
            criticality: Criticité
            nb_occurrences: Nombre d'occurrences
            
        Returns:
            Actions correctives détaillées
        """
        # Base d'actions par cause
        actions_by_cause = {
            'corrosion': [
                'Analyse chimique eau/vapeur complète',
                'Revêtement anticorrosion céramique',
                'Injection inhibiteurs corrosion',
                'Contrôle pH automatisé en continu'
            ],
            'erosion': [
                'Contrôle vitesse circulation optimisée',
                'Installation déflecteurs protection',
                'Revêtement dur anti-usure',
                'Amélioration filtration particules'
            ],
            'fissure': [
                'Inspection CND (PAUT) systématique',
                'Réparation soudage qualifié ASME',
                'Analyse contraintes éléments finis',
                'Modification supportage renforcé'
            ],
            'fatigue': [
                'Surveillance vibratoire continue IoT',
                'Renforcement supports anti-vibratoire',
                'Analyse cycles contraintes détaillée',
                'Installation amortisseurs'
            ],
            'surchauffe': [
                'Optimisation réglages combustion',
                'Capteurs température distribués',
                'Amélioration refroidissement forcé',
                'Alarmes haute température précoces'
            ]
        }
        
        # Déterminer le niveau de maintenance selon criticité ET récurrence
        if criticality <= 12 and nb_occurrences <= 2:
            maintenance_level = "Maintenance corrective planifiée"
            nb_actions = 1
        elif criticality <= 16 and nb_occurrences <= 3:
            maintenance_level = "Maintenance préventive systématique"
            nb_actions = 2
        elif criticality <= 20 or nb_occurrences <= 5:
            maintenance_level = "Maintenance préventive conditionnelle renforcée"
            nb_actions = 3
        else:
            maintenance_level = "Remise en cause conception + surveillance critique"
            nb_actions = 4
        
        # Ajustement PRIORITAIRE selon la récurrence ET multiplicité
        if nb_occurrences >= 8 or len(causes) >= 4:
            maintenance_level = "🔴 URGENCE CRITIQUE: " + maintenance_level
            nb_actions = min(nb_actions + 2, 5)
        elif nb_occurrences >= 5 or len(causes) >= 3:
            maintenance_level = "🟠 PRIORITÉ HAUTE: " + maintenance_level
            nb_actions = min(nb_actions + 1, 4)
        elif nb_occurrences >= 3 or len(causes) >= 2:
            maintenance_level = "🟡 ATTENTION: " + maintenance_level
        
        # Collecter les actions pour toutes les causes
        all_actions = set()
        for cause in causes:
            cause_lower = cause.lower()
            for cause_key, actions in actions_by_cause.items():
                if cause_key in cause_lower:
                    all_actions.update(actions[:nb_actions])
                    break
        
        # Actions par défaut si aucune correspondance
        if not all_actions:
            all_actions = {
                'Surveillance renforcée multi-paramètres',
                'Analyse expertise spécialisée approfondie',
                'Plan d\'actions correctives personnalisé',
                'Formation personnel intervention spécialisée'
            }
        
        # Limiter le nombre d'actions selon le niveau
        selected_actions = list(all_actions)[:nb_actions]
        
        # Composer le résultat final
        result = maintenance_level + " + " + " + ".join(selected_actions)
        
        # Ajouter mention spéciale pour causes multiples
        if len(causes) >= 3:
            result = f"{result} + APPROCHE GLOBALE pour {len(causes)} causes combinées"
        
        return result
    
    def _fallback_simple_grouping(self) -> pd.DataFrame:
        """
        ✅ REGROUPEMENT DE SECOURS en cas d'erreur
        """
        logger.warning("⚠️ Utilisation du regroupement de secours...")
        
        groups = []
        seen_combinations = set()
        
        for _, row in self.data.iterrows():
            try:
                # Créer une clé de regroupement (composant + sous-composant seulement)
                key = (
                    str(row.get('composant', '')).lower().strip(),
                    str(row.get('sous_composant', '')).lower().strip()
                )
                
                if key not in seen_combinations:
                    seen_combinations.add(key)
                    
                    # Trouver toutes les lignes avec cette combinaison
                    mask = (
                        (self.data['composant'].astype(str).str.lower().str.strip() == key[0]) &
                        (self.data['sous_composant'].astype(str).str.lower().str.strip() == key[1])
                    )
                    
                    subset = self.data[mask]
                    
                    # Collecter toutes les causes
                    causes_list = subset['cause'].dropna().astype(str).tolist()
                    
                    group_data = {
                        'composant': row['composant'],
                        'sous_composant': row['sous_composant'],
                        'causes_list': causes_list,
                        'nb_occurrences': len(subset),
                        'duree_moyenne': subset['duree'].mean() if 'duree' in subset.columns else 1.0,
                        'duree_max': subset['duree'].max() if 'duree' in subset.columns else 1.0,
                        'duree_min': subset['duree'].min() if 'duree' in subset.columns else 1.0,
                        'duree_totale': subset['duree'].sum() if 'duree' in subset.columns else len(subset),
                        'date_count': len(subset),
                        'line_count': len(subset)
                    }
                    
                    groups.append(group_data)
                    
            except Exception as e:
                logger.debug(f"Erreur traitement ligne regroupement secours: {e}")
                continue
        
        return pd.DataFrame(groups)
    
    # ===============================
    # MÉTHODES CONSERVÉES UTILITAIRES
    # ===============================
    
    def save_dataset_amdec(self, amdec_data: List[Dict], component: str, subcomponent: str = None) -> str:
        """✅ Sauvegarde l'AMDEC générée depuis le dataset"""
        try:
            self.amdec_df = pd.DataFrame(amdec_data)
            
            if self.amdec_df.empty:
                raise ValueError("Aucune donnée AMDEC à sauvegarder")
            
            self._clean_dataframe_for_export()
            
            output_dir = 'data/generated/amdec'
            os.makedirs(output_dir, exist_ok=True)
            
            timestamp = generate_timestamp()
            safe_component = component.lower().replace(' ', '_')
            if subcomponent:
                safe_subcomponent = subcomponent.lower().replace(' ', '_')
                filename = f"amdec_dataset_{safe_component}_{safe_subcomponent}_{timestamp}.xlsx"
            else:
                filename = f"amdec_dataset_{safe_component}_{timestamp}.xlsx"
            
            output_path = os.path.join(output_dir, filename)
            self._save_formatted_excel_CLEAN(output_path)
            
            self.output_path = output_path
            logger.info(f"✅ AMDEC dataset sauvegardée: {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"❌ Erreur sauvegarde AMDEC dataset: {e}")
            raise
    
    def generate_gammes_from_amdec(self, amdec_file_path: str = None) -> List[str]:
        """✅ Génère automatiquement les gammes depuis l'AMDEC avec images"""
        try:
            from .gamme_generator import GammeGenerator
            
            if amdec_file_path is None:
                if self.output_path is None:
                    raise ValueError("Aucun fichier AMDEC disponible")
                amdec_file_path = self.output_path
            
            if not os.path.exists(amdec_file_path):
                raise FileNotFoundError(f"Fichier AMDEC non trouvé: {amdec_file_path}")
            
            df_amdec = pd.read_excel(amdec_file_path)
            gamme_generator = GammeGenerator()
            generated_gammes = []
            
            for _, row in df_amdec.iterrows():
                try:
                    component = row['Composant'].lower().replace(' ', '_')
                    subcomponent = row['Sous-composant'].lower().replace(' ', '_')
                    criticality = int(row['C'])
                    
                    logger.info(f"🔄 Génération gamme pour {component}-{subcomponent} (C={criticality})")
                    
                    gamme_data = gamme_generator.generate(component, subcomponent, criticality)
                    
                    # Enrichir avec informations AMDEC
                    gamme_data['amdec_cause'] = row['Cause']
                    gamme_data['amdec_mode'] = row['Mode de Défaillance']
                    gamme_data['amdec_actions'] = row['Actions Correctives']
                    gamme_data['amdec_effect'] = row.get('Effet', 'Non spécifié')
                    gamme_data['amdec_function'] = row.get('Fonction', 'Non spécifié')
                    
                    output_path = gamme_generator.save_to_file(gamme_data, component, subcomponent)
                    generated_gammes.append(output_path)
                    
                    logger.info(f"✅ Gamme créée: {os.path.basename(output_path)}")
                    
                except Exception as e:
                    logger.warning(f"⚠️ Erreur génération gamme: {e}")
                    continue
            
            logger.info(f"✅ {len(generated_gammes)} gammes générées automatiquement")
            return generated_gammes
            
        except Exception as e:
            logger.error(f"❌ Erreur génération gammes: {e}")
            raise
    
    def _clean_dataframe_for_export(self):
        """✅ Nettoie le DataFrame pour export (colonnes standard uniquement)"""
        standard_columns = [
            'Composant', 'Sous-composant', 'Fonction', 'Mode de Défaillance',
            'Cause', 'Effet', 'F', 'G', 'D', 'C', 'Actions Correctives'
        ]
        
        if not self.amdec_df.empty:
            available_columns = [col for col in standard_columns if col in self.amdec_df.columns]
            self.amdec_df = self.amdec_df[available_columns]
            logger.info(f"✅ DataFrame nettoyé: {len(available_columns)} colonnes standard")
    
    def _create_clean_dataframe(self):
        """✅ Crée le DataFrame final propre"""
        if not self.amdec_data:
            standard_columns = [
                'Composant', 'Sous-composant', 'Fonction', 'Mode de Défaillance',
                'Cause', 'Effet', 'F', 'G', 'D', 'C', 'Actions Correctives'
            ]
            self.amdec_df = pd.DataFrame(columns=standard_columns)
        else:
            self.amdec_df = pd.DataFrame(self.amdec_data)
            self._clean_dataframe_for_export()
            
            if not self.amdec_df.empty:
                self.amdec_df = self.amdec_df.sort_values(
                    ['Composant', 'C'], 
                    ascending=[True, False]
                ).reset_index(drop=True)
        
        logger.info(f"✅ DataFrame PROPRE créé: {len(self.amdec_df)} entrées consolidées")
    
    def _save_formatted_excel_CLEAN(self, output_path: str):
        """✅ Sauvegarde Excel formaté proprement"""
        self._clean_dataframe_for_export()
        self.amdec_df.to_excel(output_path, index=False)
        
        try:
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            
            # Styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Formater l'en-tête
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Largeurs optimisées pour causes multiples
            column_widths = {
                'A': 20, 'B': 20, 'C': 25, 'D': 30, 'E': 25, 'F': 30,
                'G': 5, 'H': 5, 'I': 5, 'J': 5, 'K': 50
            }
            
            for i, col in enumerate(self.amdec_df.columns):
                col_letter = chr(ord('A') + i)
                if col_letter in column_widths:
                    ws.column_dimensions[col_letter].width = column_widths[col_letter]
            
            # Formatage conditionnel criticité + bordures
            for row in range(2, len(self.amdec_df) + 2):
                criticality_col = chr(ord('A') + list(self.amdec_df.columns).index('C'))
                criticality_cell = ws[f'{criticality_col}{row}']
                criticality_value = criticality_cell.value
                
                if criticality_value is not None:
                    color = get_criticality_color(criticality_value)
                    criticality_cell.fill = PatternFill(
                        start_color=color[1:], end_color=color[1:], fill_type="solid"
                    )
                
                for col in range(1, len(self.amdec_df.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            wb.save(output_path)
            logger.info(f"✅ Excel formaté sauvegardé: {output_path}")
            
        except Exception as e:
            logger.warning(f"⚠️ Erreur formatage Excel: {e}")
    
    def _load_reference_models(self) -> Dict:
        """Charge les modèles AMDEC de référence"""
        models = {}
        models_dir = 'data/models'
        
        if not os.path.exists(models_dir):
            logger.warning(f"Répertoire modèles non trouvé: {models_dir}")
            return {}
        
        for filename in os.listdir(models_dir):
            if filename.endswith('.xlsx'):
                try:
                    filepath = os.path.join(models_dir, filename)
                    df = pd.read_excel(filepath)
                    component_name = filename.replace('amdec_', '').replace('.xlsx', '')
                    models[component_name] = df
                    logger.info(f"Modèle chargé: {component_name}")
                except Exception as e:
                    logger.warning(f"Erreur chargement modèle {filename}: {e}")
        
        return models
    
    def _build_knowledge_base(self) -> Dict:
        """Construit la base de connaissances"""
        return {
            'failure_modes': {
                'corrosion': {
                    'epingle': 'Corrosion externe',
                    'collecteur': 'Corrosion interne',
                    'tube': 'Corrosion côté feu',
                    'branches': 'Corrosion sous contrainte'
                },
                'fissure': {
                    'epingle': 'Fissuration thermique',
                    'collecteur': 'Fissures intergranulaires',
                    'tube': 'Fatigue thermique',
                    'branches': 'Fissuration par contrainte'
                }
            },
            'corrective_actions': {
                'corrosion': [
                    'Revêtement céramique',
                    'Contrôle chimie eau',
                    'Injection inhibiteurs'
                ],
                'fissure': [
                    'Inspection CND systématique',
                    'Réparation soudage',
                    'Contrôle contraintes'
                ]
            }
        }
    
    def _generate_from_models(self):
        """Génère AMDEC à partir des modèles de référence"""
        logger.info("Génération AMDEC depuis modèles de référence")
        
        for component_id, component_config in ComponentConfig.COMPONENTS.items():
            for subcomp_id, subcomp_name in component_config['subcomponents'].items():
                self._generate_typical_entries(component_id, subcomp_id)
    
    def _generate_typical_entries(self, component: str, subcomponent: str):
        """Génère des entrées AMDEC typiques"""
        typical_causes = ['corrosion', 'fatigue', 'erosion']
        
        for cause in typical_causes:
            frequency = 2
            gravity = 3
            detection = 2
            criticality = calculate_criticality(frequency, gravity, detection)
            
            entry = {
                'Composant': format_component_display(component),
                'Sous-composant': format_subcomponent_display(subcomponent),
                'Fonction': 'Fonction de transfert thermique',
                'Mode de Défaillance': f'Défaillance par {cause}',
                'Cause': cause.title(),
                'Effet': 'Impact sur performance',
                'F': frequency,
                'G': gravity,
                'D': detection,
                'C': criticality,
                'Actions Correctives': 'Maintenance préventive'
            }
            
            self.amdec_data.append(entry)
    
    def _optimize_amdec(self):
        """Optimise l'AMDEC générée"""
        if not self.amdec_data:
            return
        
        # Trier par criticité décroissante
        self.amdec_data.sort(key=lambda x: x['C'], reverse=True)
        
        logger.info(f"✅ AMDEC optimisée: {len(self.amdec_data)} entrées consolidées")
    
    def save_to_file(self, output_path: str = None) -> str:
        """Sauvegarde l'AMDEC dans un fichier Excel"""
        if self.amdec_df is None:
            raise ValueError("Aucune AMDEC à sauvegarder")
        
        if output_path is None:
            os.makedirs('data/generated/amdec', exist_ok=True)
            timestamp = generate_timestamp()
            filename = f"amdec_regroupee_{timestamp}.xlsx"
            output_path = os.path.join('data/generated/amdec', filename)
        
        self._save_formatted_excel_CLEAN(output_path)
        
        self.output_path = output_path
        logger.info(f"✅ AMDEC regroupée sauvegardée: {output_path}")
        return output_path
    
    def get_statistics(self) -> Dict:
        """Retourne les statistiques de l'AMDEC"""
        if self.amdec_df is None or self.amdec_df.empty:
            return {}
        
        df = self.amdec_df
        stats = {
            'total_entries': len(df),
            'unique_components': df['Composant'].nunique(),
            'unique_subcomponents': df['Sous-composant'].nunique(),
            'avg_criticality': df['C'].mean(),
            'max_criticality': df['C'].max(),
            'criticality_distribution': {
                'negligeable': len(df[df['C'] <= 12]),
                'moyenne': len(df[(df['C'] > 12) & (df['C'] <= 16)]),
                'elevee': len(df[(df['C'] > 16) & (df['C'] <= 20)]),
                'critique': len(df[df['C'] > 20])
            }
        }
        
        # Ajouter les statistiques de regroupement
        stats.update(self.grouping_stats)
        
        return stats
    
    def get_grouping_report(self) -> Dict:
        """Retourne le rapport de regroupement"""
        return {
            'grouping_method': 'component_subcomponent_only',
            'causes_fusion': 'intelligent_concatenation',
            'frequency_calculation': 'official_grid_total_occurrences',
            'statistics': self.grouping_stats
        }